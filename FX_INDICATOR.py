# 원/달러 환율 데이터 정리
# a) 필요한 정보만 끌어와서 엑셀로 정리, 차트 만들기

import datetime
import time
start_time=time.time()

from bs4 import BeautifulSoup
headers="Mozilla/5.0"
from selenium import webdriver
options = webdriver.ChromeOptions()
options.headless = True
options.add_argument("window-size=1920x1080")
options.add_argument("user-agent="+headers)
browser=webdriver.Chrome(options=options)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.chart import Reference, LineChart
from openpyxl.chart.axis import DateAxis
import statistics

import subprocess
import pyautogui

ref_date=datetime.date(2022,1,1)
today=datetime.date.today()

# 웹스크래핑; 웹페이지 열기
url = "https://kr.investing.com/currencies/usd-krw-historical-data"
browser.get(url)
browser.find_element_by_xpath("//*[@id='flatDatePickerCanvasHol']").click()
browser.find_element_by_id("startDate").clear()
browser.find_element_by_id("startDate").send_keys(str(ref_date))
browser.find_element_by_id("endDate").clear()
browser.find_element_by_id("endDate").send_keys(str(today))
browser.find_element_by_xpath("//*[@id='applyBtn']").click()
time.sleep(2)

# 엑셀; 기본 틀
file_ref=str(ref_date).replace("-",".")
file_today=str(today).replace("-",".")
wb=Workbook()
ws_a=wb.active
ws_a.title=f"{file_ref}~{file_today}"
ws_b=wb.create_sheet("Chart")
ws_a.append(["DATE", "FX RATE", "Mean", "Median"])
ws_a.column_dimensions["A"].width=11
ws_a.column_dimensions["B"].width=9
ws_a.column_dimensions["C"].width=10
ws_a.column_dimensions["D"].width=10

# 웹스크래핑; 데이터 가져오기 / dictionary 활용
dict={}
soup = BeautifulSoup(browser.page_source, "lxml")
all_lists=soup.find("tbody").find_all("tr")
for i in range(0, len(all_lists)):
    dates=all_lists[i].find_all("td")[0].get_text().replace("년 ", "-").replace("월 ", "-").replace("일", "")
    day_fx=all_lists[i].find_all("td")[1].get_text().replace(",", "")
    dict[dates]=day_fx

# 엑셀; 데이터 입력(날짜, 환율, 산술평균, 중위값)
timedelta=today-ref_date
rows_date=timedelta.days+1
td=datetime.timedelta(days=1)
for d in range(2, rows_date+2):
    ws_a["A{}".format(d)]=ref_date
    ws_a["B{}".format(d)]=str(dict.get(str(ws_a["A{}".format(d)].value)))
    if ws_a["B{}".format(d)].value == "None":
        ws_a["B{}".format(d)] = ws_a["B{}".format(d-1)].value
        if ws_a["B{}".format(d)].value == "FX RATE":
            ws_a["B{}".format(d)]=all_lists[len(all_lists)-1].find_all("td")[1].get_text().replace(",", "")
    ref_date += td
for n in range(2, rows_date+2):
    ws_a["B{}".format(n)]=float(ws_a["B{}".format(n)].value)

MnM=[]
for n in range(2, rows_date+2):
    MnM.append(float(ws_a["B{}".format(n)].value))
for d in range(2, rows_date+2):
    ws_a["C{}".format(d)]=statistics.mean(MnM)
    ws_a["D{}".format(d)]=statistics.median(MnM)

# 엑셀; 셀 서식
for column in ws_a.columns:
    for cell in column:
        cell.alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
for row in ws_a.iter_rows():
    for cell in row:
        cell.border=thin_border
ws_a["A1"].font=Font(bold=True)
ws_a["A1"].fill = PatternFill(fgColor="ffff00", fill_type="solid")
ws_a["B1"].font=Font(bold=True)
ws_a["B1"].fill = PatternFill(fgColor="ffff00", fill_type="solid")
ws_a["C1"].font=Font(bold=True)
ws_a["D1"].font=Font(bold=True)

# 엑셀; 차트 만들기
chart_value=Reference(ws_a, min_row=1, min_col=2, max_row=rows_date+1, max_col=4)
chart=LineChart()
chart.title=f"FX RATE - {file_ref}~{file_today}"
chart.y_axis.crossAx=500
chart.x_axis=DateAxis(crossAx=100)
chart.x_axis.number_format="yyyy-mm-dd"
chart.x_axis.majorTimeUnit="days"
chart.add_data(chart_value, titles_from_data=True)

date_value = Reference(ws_a, min_col=1, min_row=2, max_row=rows_date+1)
chart.set_categories(date_value)

chart.width=39.5
chart.height=13.9
ws_b.add_chart(chart, "A1")

# 엑셀; 파일 저장
file_save=f"{file_ref} ~ {file_today} , FX(WON-DOLLAR).xlsx"
wb.save(file_save)
subprocess.Popen([file_save], shell=True)
time.sleep(2)
pyautogui.hotkey("ctrl", "pgdn")
pyautogui.hotkey("ctrl", "s")
end_time=time.time()
time_spent = end_time - start_time
print(f"=> FIN. ({str(round((time_spent), 4))} sec.)")