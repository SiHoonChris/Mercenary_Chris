# 단순이동평균선(SMA-L) - 당일 포함 최근 300일
# 단순이동평균선 상단밴드(SMA-UL) - SMA-L의 +1% 값을 연결한 선
# 단순이동평균선 하단밴드(SMA-LL) - SMA-L의 -1% 값을 연결한 선
# End-Price가 SMA-UL 위에 있으면 'Sell'
# End-Price가 SMA-LL 아래 있으면 'Buy'
# End-Price가 SMA-UL과 SMA-LL 사이에 있으면 'Hold'
# Median - 당일 포함 최근 100일에 대한 중위값
# Mean - 당일 포함 최근 100일에 대한 산술평균값
# 변동성(Vlt. ; Volatility) -  " Vlt. = |Median - Mean| "
# Column Title : Date,  End-Price,  SMA-L,  SMA-UL,  SMA-LL,  Buy-Sell-Hold,  Median, Mean, Vlt.


from openpyxl import load_workbook
from datetime import datetime
import pandas as pd


# 데이터 가공
file='USD_KRW 내역 (19.8.21~22.8.21).xlsx'   # 출처 : kr.investing.com
wb = load_workbook(file)
ws = wb.active

for i in range(2, ws.max_row+1):
    cell=ws['A{}'.format(i)].value
    dt = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + cell - 2)
    ws['A{}'.format(i)]=dt

wb.save(file)
# 'pandas로 excel파일 끌어올 때 날짜 데이터가 5자리 숫자로 가져와짐' => 해결
# stackoverflow 찾아봄


# 데이터 가공 : 기본 프레임
df = pd.read_excel(file)
df.sort_index(ascending=False, inplace=True)
df.reset_index(drop=True, inplace=True)

print(df)