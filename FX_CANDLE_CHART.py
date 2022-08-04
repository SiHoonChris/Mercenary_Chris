# 원/달러 환율 데이터 정리
# b) Price-values( End(Close) / Open / High / Low ) 활용하여 캔들차트 만들기 - 일, 주, 월봉에 대한 캔들차트

import pandas as pd

from bs4 import BeautifulSoup
headers="Mozilla/5.0"
from selenium import webdriver
options = webdriver.ChromeOptions()
options.headless = True
options.add_argument("window-size=1920x1080")
options.add_argument("user-agent="+headers)
browser=webdriver.Chrome(options=options)

import datetime
import time
from openpyxl import Workbook
import subprocess
import pyautogui
start_time=time.time()

ref_date=datetime.date(2022,1,1)
today=datetime.date.today()

# 웹스크래핑; 웹페이지 열기
url = "https://kr.investing.com/currencies/usd-krw-historical-data"
browser.get(url)
browser.find_element_by_xpath("//*[@id='flatDatePickerCanvasHol']").click()
browser.find_element_by_id("startDate").clear()
browser.find_element_by_id("startDate").send_keys(str(ref_date))
browser.find_element_by_id("endDate").clear()
browser.find_element_by_id("endDate").send_keys(str(today))
browser.find_element_by_xpath("//*[@id='applyBtn']").click()
time.sleep(2)

# 엑셀; 기본 틀
wb=Workbook()
ws=wb.active
file_ref=str(ref_date).replace("-", ".")
file_today=str(today).replace("-", ".")
ws.title=f"{file_ref}~{file_today}"
ws.append(["Date", "End", "Open", "Highest", "Lowest"])
ws.column_dimensions["A"].width=11
ws.column_dimensions["B"].width=9
ws.column_dimensions["C"].width=9
ws.column_dimensions["D"].width=9
ws.column_dimensions["E"].width=9

# 웹스크래핑; 데이터 가져오기(날짜, 환율 - 종가, 오픈, high, low)
soup = BeautifulSoup(browser.page_source, "lxml")
all_lists=soup.find("tbody").find_all("tr")
for i in range(0, len(all_lists)):
    Dates=all_lists[i].find_all("td")[0].get_text().replace("년 ", "-").replace("월 ", "-").replace("일", "")
    Ends=all_lists[i].find_all("td")[1].get_text().replace(",", "")
    Opens=all_lists[i].find_all("td")[2].get_text().replace(",", "")
    Highests=all_lists[i].find_all("td")[3].get_text().replace(",", "")
    Lowests=all_lists[i].find_all("td")[4].get_text().replace(",", "")
    ws["A{}".format(i+2)]=Dates
    ws["B{}".format(i+2)]=float(Ends)
    ws["C{}".format(i+2)]=float(Opens)
    ws["D{}".format(i+2)]=float(Highests)
    ws["E{}".format(i+2)]=float(Lowests)

# 엑셀; 파일 저장
file_save=f"{file_ref} ~ {file_today}, FX(WON-DOLLAR), Candle.xlsx"
wb.save(file_save)
subprocess.Popen([file_save], shell=True)
time.sleep(2)
pyautogui.hotkey("ctrl", "s")
end_time=time.time()
time_spent = end_time - start_time
print(f"=> FIN. ({str(round((time_spent), 4))} sec.)")

# excel자료 DataFrame화 시키기 & 오름차순 정렬
df = pd.read_excel(file_save)
df.set_index('Date', inplace=True)
df.sort_index(ascending=False)
print(df)
