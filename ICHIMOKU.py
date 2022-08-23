import subprocess
import pyautogui
import time
from openpyxl.styles import Border, Side,  PatternFill
from openpyxl import load_workbook
import pandas as pd

# 데이터 가공 : 기본 프레임
df = pd.read_csv('IVV, 21.01.01~21.12.31.csv')
df.sort_index(ascending=False, inplace=True)
df.reset_index(drop=True, inplace=True)

df.drop(columns=['오픈', '고가', '저가', '거래량', '변동 %'], inplace=True)
df['전환선']=0
df['기준선']=0
df['선행스팬A(1)']=0
df['선행스팬B(2)']=0
df['구름대']=''
df['추세']=''

last_idx=df.index[-1]
for i in range(1, 26):
    df.loc[last_idx+i] = ['', 0, '', '', 0, 0, '', '미정']

# 전환선
for i in df.index:
    if i < 9:
        df.loc[i, '전환선']=''
    elif 9 <= i <= last_idx:
        max_value=max(df.loc[i-9:i-1 , '종가'])
        min_value=min(df.loc[i-9:i-1 , '종가'])
        df.loc[i, '전환선']=(max_value+min_value)/2
    else:
        df.loc[i, '전환선']=''
        
# 기준선
for i in df.index:
    if i < 26:
        df.loc[i, '기준선']=''
    elif 26 <= i <= last_idx:
        max_value=max(df.loc[i-26:i-1 , '종가'])
        min_value=min(df.loc[i-26:i-1 , '종가'])
        df.loc[i, '기준선']=(max_value+min_value)/2
    else:
        df.loc[i, '기준선']=''

# 선행스팬A(1)
for i in df.index:
    if i < 26:
        df.loc[i, '선행스팬A(1)']=0
        df.loc[i+25, '선행스팬A(1)']=0
    elif 26 <= i <= last_idx:
        df.loc[i+25, '선행스팬A(1)']=(df.loc[i, '전환선']+df.loc[i, '기준선'])/2
    else:
        pass
    
# 선행스팬B(2)
for i in df.index:
    if i < 51:
        df.loc[i, '선행스팬B(2)']=0
        df.loc[i+25, '선행스팬B(2)']=0
    elif 51 <= i <= last_idx:
        max_value=max(df.loc[i-51:i , '종가'])
        min_value=min(df.loc[i-51:i , '종가'])
        df.loc[i+25, '선행스팬B(2)']=(max_value+min_value)/2
    else:
        pass
    
# 구름대
for i in df.index:
    if df.loc[i, '선행스팬B(2)']==0:
        pass
    else:
        if df.loc[i, '선행스팬A(1)'] > df.loc[i, '선행스팬B(2)']:
            df.loc[i, '구름대'] = '양운'
        elif df.loc[i, '선행스팬A(1)'] < df.loc[i, '선행스팬B(2)']:
            df.loc[i, '구름대'] = '음운'
        else:
            df.loc[i, '구름대'] = '구름선'
            
# 추세
for i in df.index:
    if df.loc[i, '선행스팬B(2)']==0 or df.loc[i, '종가']==0:
        pass
    else:
        if df.loc[i, '종가'] > df.loc[i, '선행스팬A(1)'] and df.loc[i, '종가'] > df.loc[i, '선행스팬B(2)']:
            df.loc[i, '추세'] = '상승'
        elif df.loc[i, '종가'] < df.loc[i, '선행스팬A(1)'] and df.loc[i, '종가'] < df.loc[i, '선행스팬B(2)']:
            df.loc[i, '추세'] = '하락'
        else:
            df.loc[i, '추세'] = '전환'
            
# 데이터 가공 : 정리
for i in range(1, 26):
    df.loc[last_idx+i, '종가'] = ''
for i in df.index:
    if df.loc[i, '선행스팬A(1)']==0:
        df.loc[i, '선행스팬A(1)']=''
    if df.loc[i, '선행스팬B(2)']==0:
        df.loc[i, '선행스팬B(2)']=''

# DataFrame 저장
print(df)
title='일목균형표(IVV, 21.01.01~21.12.31).xlsx'
df.to_excel(title)


# 엑셀 작업
wb = load_workbook(title)
ws = wb.active

# 전체 셀에 대해 라인 구분
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
for row in ws.iter_rows():
    for cell in row:
        cell.border=thin_border

# 인덱스 Column 삭제
ws.delete_cols(1)

# 셀 너비 조정(날짜, 선행스팬1(A), 선행스팬2(B))
ws.column_dimensions["A"].width=16
ws.column_dimensions["E"].width=11
ws.column_dimensions["F"].width=11

# 구름대, 색 구분
# 양운:빨강, 음운:파랑, 구름선:노랑, '값 없음':'색 없음'
for i in range(1, ws.max_row+1):
    if ws["G{}".format(i)].value=='양운':
        ws["G{}".format(i)].fill = PatternFill(fgColor="ff704d", fill_type="solid")
    elif ws["G{}".format(i)].value=='음운':
        ws["G{}".format(i)].fill = PatternFill(fgColor="6699ff", fill_type="solid")
    elif ws["G{}".format(i)].value=='구름선':
        ws["G{}".format(i)].fill = PatternFill(fgColor="ffff99", fill_type="solid")
    else:
        pass

# 추세, 색 구분
# 상승:빨강, 하락:파랑, 전환:노랑, 미정:회색(#F2F2F2), '값 없음':'색 없음'
for i in range(1, ws.max_row+1):
    if ws["H{}".format(i)].value=='상승':
        ws["H{}".format(i)].fill = PatternFill(fgColor="ff704d", fill_type="solid")
    elif ws["H{}".format(i)].value=='하락':
        ws["H{}".format(i)].fill = PatternFill(fgColor="6699ff", fill_type="solid")
    elif ws["H{}".format(i)].value=='전환':
        ws["H{}".format(i)].fill = PatternFill(fgColor="ffff99", fill_type="solid")
    elif ws["H{}".format(i)].value=='미정':
        ws["H{}".format(i)].fill = PatternFill(fgColor="cccccc", fill_type="solid")
    else:
        pass

# 틀 고정
ws.freeze_panes = "B2"

# 저장 & 바로보기
wb.save(title)
subprocess.Popen([title], shell=True)
time.sleep(2)
pyautogui.hotkey("ctrl", "s")